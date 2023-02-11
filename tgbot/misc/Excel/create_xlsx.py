import openpyxl
from openpyxl.styles import Alignment

from tgbot.services.database import db_commands
from tgbot.services.database.db_models import Questionnaire
from tgbot.services.service_functions import get_average_completion_time


async def create_xlsx_file(questionnaire: Questionnaire):
    book = openpyxl.Workbook()
    book.remove(book.active)
    sheet_1 = book.create_sheet("Ответы")
    sheet_2 = book.create_sheet("Статистика")
    qe_title = questionnaire.title

    questions_fields = await db_commands.select_questions(qe_id=questionnaire.qe_id)
    questions = []
    for field in questions_fields:
        questions.append(field.question_text)
    questions.append("Контактная информация")
    sheet_1.append(questions)

    passed_qes = await db_commands.select_passed_users(qe_id=questionnaire.qe_id)
    for passed_qe in passed_qes:
        answers_fields = await db_commands.select_user_answers(respondent_id=passed_qe.respondent_id,
                                                               qe_id=questionnaire.qe_id)
        answers = []
        for field in answers_fields:
            answers.append(field.answer_text)
        respondent = await db_commands.select_user(id=passed_qe.respondent_id)
        answers.append(respondent.email)
        sheet_1.append(answers)

    letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
    i = 0
    for letter in letters:
        if i > questionnaire.questions_quantity:
            break
        col = sheet_1.column_dimensions[f'{letter}']
        col.width = len(questions[i]) + 10
        col.alignment = Alignment(horizontal="center")
        i += 1

    if questionnaire.passed_by == 0:
        pass_percent = 0
    else:
        pass_percent = questionnaire.passed_by / questionnaire.started_by * 100

    average_ct = await get_average_completion_time(qe_id=questionnaire.qe_id)

    sheet_2.append(["📊 Статистика опроса:"])
    sheet_2.append([f"• Начали проходить: {questionnaire.started_by} чел."])
    sheet_2.append([f"• Прошли: {questionnaire.passed_by} чел."])
    sheet_2.append([f"• Процент прохождения: {pass_percent:.1f}%"])
    sheet_2.append([f"• Среднее время прохождения: {average_ct[0]:.1f} {average_ct[1]}"])
    sheet_2.append([f"• Дата создания: {questionnaire.created_at}"])

    col = sheet_2.column_dimensions[f'A']
    col.width = 50

    path = rf'C:\Users\User\PycharmProjects\Questionnaire_UniBot\tgbot\misc\Excel\xlsx_files\{qe_title}.xlsx'
    book.save(path)

    return path
